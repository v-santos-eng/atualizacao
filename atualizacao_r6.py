import os
import pandas as pd
import time
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import tkinter as tk
from tkinter import ttk
from threading import Thread

# Diretórios fixos
DIRETORIOS_ORIGEM = [
    r"\\DCSCHVM168\TransferCPI\salesOrder\uploaded\BR",
    r"\\DCSCHVM168\TransferCPI\salesQuotation\uploaded\BR"
]
DIRETORIO_DESTINO = r"Y:\BR10\IndustrialEngineering\Documents\KPI\python BI\base de dados"
ARQUIVO_DESTINO = "base_de_dados.xlsx"

def listar_arquivos(diretorios, extensao=".csv"):
    """Lista arquivos com a extensão especificada em múltiplos diretórios."""
    arquivos = []
    for diretorio in diretorios:
        try:
            arquivos += [os.path.join(diretorio, f) for f in os.listdir(diretorio) if f.endswith(extensao)]
        except Exception as e:
            print(f"Erro ao acessar o diretório {diretorio}: {e}")
    return arquivos

def processar_csv(caminho_arquivo):
    """
    Processa o arquivo .csv para verificar a primeira célula e determinar a aba correta.
    Retorna o tipo (Sales Orders ou Quotation), DataFrame do CSV, o número extraído e a data de criação.
    """
    try:
        df = pd.read_csv(caminho_arquivo, header=None, delimiter=";", encoding='latin1')
        primeira_celula = str(df.iloc[0, 0]).strip()
        if primeira_celula.startswith('5') or primeira_celula.startswith('2'):
            numero = primeira_celula.strip()
            data_criacao = datetime.fromtimestamp(os.path.getctime(caminho_arquivo)).strftime("%d/%m/%Y")
            return ('Sales Orders' if numero.startswith('5') else 'Quotations'), df, numero, data_criacao, None
        else:
            return None, None, None, None, None
    except Exception as e:
        print(f"Erro ao processar o arquivo {caminho_arquivo}: {e}")
        return None, None, None, None, None

def extrair_cliente(nome_arquivo):
    """
    Extrai o nome do cliente a partir do nome do arquivo.
    Assume que o nome do cliente é a última parte após o último '_'.
    """
    try:
        cliente = nome_arquivo.split('_')[-1].replace(".csv", "").strip()
        return cliente
    except IndexError:
        return "Cliente Desconhecido"

def extrair_dados(df, tipo, numero, data_criacao, customer):
    """
    Extrai os dados dos códigos que começam com 'AT-' e suas quantidades.
    Inverte os dados: o 'AT-' vai para a coluna 'Module', e o número vai para 'Sales Order' ou 'Quotation'.
    """
    dados = []
    for i, row in df.iterrows():
        codigo = str(row[2]).strip() if len(row) > 2 else None # Terceira coluna (índice 2)
        if codigo and codigo.startswith("AT-"):
            quantidade = int(row[3]) if len(row) > 3 and str(row[3]).isdigit() else None
            comprimento = buscar_comprimento(df, i, codigo)
            between_frames = buscar_informacao(df, i, "AT_MCP2_BTF_LEN_M")
            pitch = buscar_informacao(df, i, "AT_MCP2_PIT_M")
            gear_ratio = buscar_informacao(df, i, "AT_MCP2_GEA_01_RAT")
            power = buscar_informacao(df, i, "AT_MCP2_RD_POW_M")
            voltage = buscar_informacao(df, i, "AT_MCP2_RD_VOT")
            interface_type = buscar_informacao(df, i, "AT_MCP2_RD_INT_TYP")
            sensor_type = buscar_informacao(df, i, "AT_MCP2_SEN_SPLR_01")
            tor = buscar_informacao(df, i, "AT_MCP2_TOP_LVL_CNV_HEI_M") # Coluna TOR
            control_card = buscar_informacao(df, i, "AT_MCP2_CTR_CRD_TYP") # Coluna Control Card
            zone_length = buscar_informacao(df, i, "AT_MCP2_ZON_LEN_M") # Coluna Zone Length
            electric_side = buscar_informacao(df, i, "AT_MCP2_ELC_SID") # Coluna Electric Side
            side_guide_left = buscar_informacao(df, i, "AT_MCP2_SGD_LFT_TYP") # Coluna Side Guide Left Type
            side_guide_right = buscar_informacao(df, i, "AT_MCP2_SGD_RGT_TYP") # Coluna Side Guide Right Type
            bus_type = buscar_informacao(df, i, "AT_MCP2_CTR_BUS_TYP") # Coluna Bus Type
            msc_quantity = buscar_informacao(df, i, "AT_MCP2_ROL_MSC_01_QTY") # Coluna MSC Quantity
            merge_divert = buscar_informacao(df, i, "AT_MCP2_MRG_DIV_SEL") # Coluna Merge/Divert
            merge_divert_angle = buscar_informacao(df, i, "AT_MCP2_MOD_MRG_ANG") # Coluna Merge/Divert Angle
            alignment_angle = buscar_informacao(df, i, "AT_MCP2_FKT_ANG") # Coluna Alignment Angle
            motor_position = buscar_informacao(df, i, "AT_MCP2_DRV_UNT_POS") # Coluna Motor Position
            motor_manufacturer = buscar_informacao(df, i, "AT_MCP2_MOT_MNF") # Coluna Motor Manufacturer
            framebed_type = buscar_informacao(df, i, "AT_MCP2_FRB_TYP") # Coluna Framebed Type
            sword_quantity = buscar_informacao(df, i, "AT_MCP2_TRF_SWO_QTY") # Coluna Sword Quantity
            cassetes_quantity = buscar_informacao(df, i, "AT_MCP2_CAS_QTY") # Coluna Cassetes Quantity
            lower_conveyor_height = buscar_informacao(df, i, "AT_MCP2_LOW_LVL_CNV_HEI_M") # Coluna Lower Conveyor Height
            higher_conveyor_height = buscar_informacao(df, i, "AT_MCP2_TOP_LVL_CNV_HEI_M") # Coluna Higher Conveyor Height
            support_type = buscar_informacao(df, i, "AT_MCP2_SP_TYP_01") # Coluna Support Type
            delivery_date = str(row[8]).strip() if len(row) > 8 else ""
            if delivery_date:
                delivery_date = delivery_date.replace('.', '/')
            dados.append([numero, codigo, quantidade, comprimento, between_frames, pitch, gear_ratio, power, voltage,
                          interface_type, sensor_type, tor, control_card, zone_length, electric_side, side_guide_left,
                          side_guide_right, bus_type, msc_quantity, merge_divert, merge_divert_angle, alignment_angle,
                          motor_position, motor_manufacturer, framebed_type, sword_quantity, cassetes_quantity,
                          lower_conveyor_height, higher_conveyor_height, support_type, data_criacao, delivery_date, customer])
    return dados

def buscar_comprimento(df, index, codigo):
    """
    Busca o comprimento relacionado a 'AT_MCP2_MOD_LEN_M' ou aplica regra especial para 'AT-RM8320-E2'.
    """
    try:
        if codigo == "AT-RM8320-E2":
            for i in range(index + 1, len(df)): # Busca nas linhas seguintes
                row = df.iloc[i]
                for col in range(len(row)):
                    if "AT_MCP2_FKT_ANG" in str(row[col]):
                        angulo = row[col + 1] if col + 1 < len(row) else None
                        if angulo in [30, 45, 60, 90]:
                            return {30: 542, 45: 813, 60: 1084, 90: 1626}.get(angulo, 0)
        for i in range(index + 1, len(df)): # Busca nas linhas seguintes
            row = df.iloc[i]
            for col in range(len(row)):
                if "AT_MCP2_MOD_LEN_M" in str(row[col]):
                    valor = row[col + 1] if col + 1 < len(row) else None
                    return int(float(valor)) if str(valor).replace('.', '', 1).isdigit() else 0
    except Exception as e:
        print(f"Erro ao buscar comprimento: {e}")
    return 0
def buscar_informacao(df, index, chave):
    """
    Busca uma informação específica pela chave no arquivo CSV.
    Se não encontrada, retorna 0.
    """
    try:
        for i in range(index + 1, len(df)): # Busca nas linhas seguintes
            row = df.iloc[i]
            for col in range(len(row)):
                if chave in str(row[col]):
                    valor = row[col + 1] if col + 1 < len(row) else None
                    # Substitui "NA" por "None"
                    if valor == "NA":
                        return "None"
                    return str(valor).strip() if valor else 0
    except Exception as e:
        print(f"Erro ao buscar a informação '{chave}': {e}")
    return 0

def ajustar_coluna(ws, start_row):
    """
    Ajusta a primeira coluna da aba removendo os dois últimos caracteres a partir de uma linha específica.
    """
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=1): # Apenas coluna 1 (A)
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value[:-2] # Remove os dois últimos caracteres

def formatar_como_tabela(ws, nome_tabela):
    """
    Formata os dados da planilha como tabela.
    """
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tabela = Table(displayName=nome_tabela, ref=ref)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)

def atualizar_tabela_existente(ws, dados):
    """
    Atualiza a tabela existente com novos dados.
    """
    start_row = ws.max_row + 1
    for item in dados:
        ws.append(item)
    ajustar_coluna(ws, start_row)
    # Atualizar a referência da tabela existente
    for tabela in ws.tables.values():
        tabela.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

def salvar_em_excel(sales_orders_data, quotations_data, caminho_destino, nome_arquivo):
    """
    Salva os dados nas abas 'Sales Orders' e 'Quotations' em um arquivo Excel.
    """
    try:
        caminho_completo = os.path.join(caminho_destino, nome_arquivo)
        if os.path.exists(caminho_completo):
            wb = load_workbook(caminho_completo)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # Remove a aba padrão criada automaticamente

        # Aba Sales Orders
        if "Sales Orders" in wb.sheetnames:
            sales_orders_ws = wb["Sales Orders"]
            atualizar_tabela_existente(sales_orders_ws, sales_orders_data)
        else:
            sales_orders_ws = wb.create_sheet("Sales Orders")
            sales_orders_ws.append(["Sales Order", "Module", "Quantity", "Length", "Between Frames", "Pitch", "Gear Ratio", 
                                    "Power (W)", "Voltage (V)", "Interface Type", "Sensor Type", "TOR", "Control Card", 
                                    "Zone Length", "Eletric Side", "Side Guide Left Type", "Side Guide Right Type", 
                                    "Bus Type", "MSC Quantity", "Merge/Divert", "Merge/Divert Angle", "Alignment Angle", 
                                    "Motor Position", "Motor Manufacturer", "Framebed Type", "Sword Quantity", 
                                    "Cassetes Quantity", "Lower Conveyor Height (TOR1)", "Higher Conveyor Height (TOR21)", 
                                    "Support Type", "Creation Date", "Delivery Date", "Customer"])
            for item in sales_orders_data:
                sales_orders_ws.append(item)
            # Formatar como tabela
            formatar_como_tabela(sales_orders_ws, "SalesOrdersTable")

        # Aba Quotations
        if "Quotations" in wb.sheetnames:
            quotations_ws = wb["Quotations"]
            atualizar_tabela_existente(quotations_ws, quotations_data)
        else:
            quotations_ws = wb.create_sheet("Quotations")
            quotations_ws.append(["Quotation", "Module", "Quantity", "Module Length", "Between Frames", "Pitch", "Gear Ratio", 
                                  "Power (W)", "Voltage (V)", "Interface Type", "Sensor Type", "TOR", "Control Card", 
                                  "Zone Length", "Eletric Side", "Side Guide Left Type", "Side Guide Right Type", 
                                  "Bus Type", "MSC Quantity", "Merge/Divert", "Merge/Divert Angle", "Alignment Angle", 
                                  "Motor Position", "Motor Manufacturer", "Framebed Type", "Sword Quantity", 
                                  "Cassetes Quantity", "Lower Conveyor Height (TOR1)", "Higher Conveyor Height (TOR21)", 
                                  "Support Type", "Creation Date", "Delivery Date", "Customer"])
            for item in quotations_data:
                quotations_ws.append(item)
            # Formatar como tabela
            formatar_como_tabela(quotations_ws, "QuotationsTable")

        wb.save(caminho_completo)
        print(f"Arquivo Excel salvo com sucesso em: {caminho_completo}")
    except Exception as e:
        print(f"Erro ao salvar o arquivo Excel: {e}")

def ler_numeros_excel(arquivo_excel):
    # Ler o arquivo Excel
    xls = pd.ExcelFile(arquivo_excel)
    # Ler a primeira coluna da aba Sales Orders
    sales_orders = pd.read_excel(xls, 'Sales Orders', usecols=[0], engine='openpyxl')
    sales_orders_numeros = sales_orders.iloc[:, 0].drop_duplicates().astype(str).tolist()
    # Ler a primeira coluna da aba Quotations
    quotations = pd.read_excel(xls, 'Quotations', usecols=[0], engine='openpyxl')
    quotations_numeros = quotations.iloc[:, 0].drop_duplicates().astype(str).tolist()
    return sales_orders_numeros, quotations_numeros

def buscar_arquivos(diretorio, numeros):
    arquivos_nao_correspondentes = []
    for arquivo in os.listdir(diretorio):
        caminho_completo = os.path.join(diretorio, arquivo)
        if os.path.isfile(caminho_completo):
            encontrado = False
            for numero in numeros:
                if numero in arquivo:
                    encontrado = True
                    break
            if not encontrado:
                arquivos_nao_correspondentes.append(arquivo)
    return arquivos_nao_correspondentes

def gerar_log_entry(entry_type, message):
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    return f"{timestamp} - {entry_type}: {message}\n"

def listar_arquivos_e_verificar(diretorios, arquivo_excel, log_file):
    sales_orders_numeros, quotations_numeros = ler_numeros_excel(arquivo_excel)
    with open(log_file, 'a') as log: # Abrir o arquivo de log no modo de adição
        log.write(gerar_log_entry("INFO", "Iniciando verificação de arquivos"))
        for diretorio in diretorios:
            log.write(gerar_log_entry("INFO", f"Verificando arquivos no diretório {diretorio}"))
            # Buscar arquivos no diretório para Sales Orders e Quotations
            arquivos_nao_correspondentes_sales_orders = buscar_arquivos(diretorio, sales_orders_numeros)
            arquivos_nao_correspondentes_quotations = buscar_arquivos(diretorio, quotations_numeros)
            # Combinar listas de arquivos não correspondentes e remover duplicatas
            arquivos_nao_correspondentes = list(set(arquivos_nao_correspondentes_sales_orders + arquivos_nao_correspondentes_quotations))
            if arquivos_nao_correspondentes:
                log.write(gerar_log_entry("ERROR", f"Arquivos não correspondentes encontrados: {arquivos_nao_correspondentes}"))
            else:
                log.write(gerar_log_entry("INFO", "Todos os arquivos correspondem aos números das planilhas"))
        log.write("\n")

def atualizar_barra_progresso(progress, passo_atual, total_passos, descricao, log, arquivos_lidos, total_arquivos, tempo_inicio):
    """
    Atualiza a barra de progresso, a descrição do passo atual, o log de progresso e o tempo estimado restante.
    """
    progress['value'] = (passo_atual / total_passos) * 100
    progress.update_idletasks()
    descricao.set(f"Passo {passo_atual} de {total_passos}: {descricao}")
    log.set(f"Arquivos lidos: {arquivos_lidos} de {total_arquivos}")
    # Calcular tempo decorrido
    tempo_decorrido = time.time() - tempo_inicio
    minutos_decorridos, segundos_decorridos = divmod(tempo_decorrido, 60)
    tempo_decorrido_str = f"Tempo decorrido: {int(minutos_decorridos)}m {int(segundos_decorridos)}s"
    # Calcular tempo estimado restante
    if arquivos_lidos > 0:
        tempo_por_arquivo = tempo_decorrido / arquivos_lidos
        tempo_restante = tempo_por_arquivo * (total_arquivos - arquivos_lidos)
        minutos_restantes, segundos_restantes = divmod(tempo_restante, 60)
        tempo_restante_str = f"Tempo estimado restante: {int(minutos_restantes)}m {int(segundos_restantes)}s"
    else:
        tempo_restante_str = "Tempo estimado restante: calculando..."
    log.set(f"Arquivos lidos: {arquivos_lidos} de {total_arquivos}\n{tempo_decorrido_str}\n{tempo_restante_str}")

def executar_script():
    inicio = time.time() # Marca o inicio do programa
    # Listar arquivos
    descricao.set("Listando arquivos...")
    root.update()
    arquivos_csv = listar_arquivos(DIRETORIOS_ORIGEM)
    atualizar_barra_progresso(progress, 1, total_passos, descricao, log, 0, len(arquivos_csv), inicio)
    if arquivos_csv:
        sales_orders_data = []
        quotations_data = []
        # Processar arquivos CSV
        descricao.set("Processando arquivos CSV...")
        root.update()
        arquivos_lidos = 0
        for arquivo in arquivos_csv:
            caminho_arquivo = arquivo
            customer = extrair_cliente(os.path.basename(arquivo))
            tipo, df, numero, data_criacao, _ = processar_csv(caminho_arquivo)
            if tipo and df is not None:
                dados = extrair_dados(df, tipo, numero, data_criacao, customer)
                if tipo == 'Sales Orders':
                    sales_orders_data.extend(dados)
                elif tipo == 'Quotations':
                    quotations_data.extend(dados)
                arquivos_lidos += 1
                atualizar_barra_progresso(progress, 2, total_passos, descricao, log, arquivos_lidos, len(arquivos_csv), inicio)
        # Salvar em Excel
        descricao.set("Salvando em Excel...")
        root.update()
        salvar_em_excel(sales_orders_data, quotations_data, DIRETORIO_DESTINO, ARQUIVO_DESTINO)
        atualizar_barra_progresso(progress, 3, total_passos, descricao, log, arquivos_lidos, len(arquivos_csv), inicio)
    else:
        print("Nenhum arquivo CSV encontrado nos diretórios de origem.")
    fim = time.time() # Marca o final do programa
    tempo_total = fim - inicio
    print(f"Tempo total de execução: {tempo_total:.2f} segundos")
    descricao.set("Concluído!")
    root.update()

if __name__ == "__main__":
    # Configuração da janela de progresso
    root = tk.Tk()
    root.title("Progresso do Script")
    root.geometry("400x150")
    descricao = tk.StringVar()
    descricao.set("Iniciando...")
    log = tk.StringVar()
    log.set("")
    progress = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
    progress.pack(pady=10)
    label_descricao = tk.Label(root, textvariable=descricao)
    label_descricao.pack()
    label_log = tk.Label(root, textvariable=log)
    label_log.pack()
    total_passos = 3
    # Iniciar a execução do script em uma thread separada
    thread = Thread(target=executar_script)
    thread.start()
    # Manter a janela de progresso aberta
    root.mainloop()