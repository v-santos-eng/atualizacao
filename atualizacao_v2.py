import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Diretórios fixos
DIRETORIOS_ORIGEM = [
    r"C:\Users\v.santos\OneDrive - Interroll Management SA\Desktop\projeto python",
    #r"\\DCSCHVM168\\TransferCPI\\salesQuotation\\uploaded\\BR"
]
DIRETORIO_DESTINO = r"C:\Users\v.santos\OneDrive - Interroll Management SA\Desktop\projeto python\atualizacao"
ARQUIVO_DESTINO = "base_de_dados.xlsx"

def listar_arquivos(diretorios, extensao=".csv"):
    """Lista arquivos com a extensão especificada em múltiplos diretórios."""
    arquivos = []
    for diretorio in diretorios:
        try:
            arquivos += [os.path.join(diretorio, f) for f in os.listdir(diretorio) if f.endswith(extensao)]
        except Exception as e:
            print(f"Erro ao acessar o diretório {diretorio}: {e}")
    return arquivos

def processar_csv(caminho_arquivo):
    """
    Processa o arquivo .csv para verificar a primeira célula e determinar a aba correta.
    Retorna o tipo (Sales Orders ou Quotation), DataFrame do CSV, o número extraído e a data de criação.
    """
    try:
        df = pd.read_csv(caminho_arquivo, header=None, delimiter=";")
        primeira_celula = str(df.iloc[0, 0]).strip()
        if primeira_celula.startswith('5') or primeira_celula.startswith('2'):
            numero = primeira_celula.strip()
            data_criacao = datetime.fromtimestamp(os.path.getctime(caminho_arquivo)).strftime("%d/%m/%Y")
            return ('Sales Orders' if numero.startswith('5') else 'Quotations'), df, numero, data_criacao, None
        else:
            return None, None, None, None, None
    except Exception as e:
        print(f"Erro ao processar o arquivo {caminho_arquivo}: {e}")
        return None, None, None, None, None

def extrair_cliente(nome_arquivo):
    """
    Extrai o nome do cliente a partir do nome do arquivo.
    Assume que o nome do cliente é a última parte após o último '_'.
    """
    try:
        cliente = nome_arquivo.split('_')[-1].replace(".csv", "").strip()
        return cliente
    except IndexError:
        return "Cliente Desconhecido"

def extrair_dados(df, tipo, numero, data_criacao, customer):
    """
    Extrai os dados dos códigos que começam com 'AT-' e suas quantidades.
    Inverte os dados: o 'AT-' vai para a coluna 'Module', e o número vai para 'Sales Order' ou 'Quotation'.
    """
    dados = []
    for i, row in df.iterrows():
        codigo = str(row[2]).strip() if len(row) > 2 else None  # Terceira coluna (índice 2)
        if codigo and codigo.startswith("AT-"):
            quantidade = int(row[3]) if len(row) > 3 and str(row[3]).isdigit() else None
            comprimento = buscar_comprimento(df, i, codigo)
            between_frames = buscar_informacao(df, i, "AT_MCP2_BTF_LEN_M")
            pitch = buscar_informacao(df, i, "AT_MCP2_PIT_M")
            gear_ratio = buscar_informacao(df, i, "AT_MCP2_GEA_01_RAT")
            power = buscar_informacao(df, i, "AT_MCP2_RD_POW_M")
            voltage = buscar_informacao(df, i, "AT_MCP2_RD_VOT")
            interface_type = buscar_informacao(df, i, "AT_MCP2_RD_INT_TYP")
            sensor_type = buscar_informacao(df, i, "AT_MCP2_SEN_SPLR_01")
            tor = buscar_informacao(df, i, "AT_MCP2_TOP_LVL_CNV_HEI_M")  # Coluna TOR
            control_card = buscar_informacao(df, i, "AT_MCP2_CTR_CRD_TYP")  # Coluna Control Card
            zone_length = buscar_informacao(df, i, "AT_MCP2_ZON_LEN_M")  # Coluna Zone Length
            electric_side = buscar_informacao(df, i, "AT_MCP2_ELC_SID")  # Coluna Electric Side
            side_guide_left = buscar_informacao(df, i, "AT_MCP2_SGD_LFT_TYP")  # Coluna Side Guide Left Type
            side_guide_right = buscar_informacao(df, i, "AT_MCP2_SGD_RGT_TYP")  # Coluna Side Guide Right Type
            bus_type = buscar_informacao(df, i, "AT_MCP2_CTR_BUS_TYP")  # Coluna Bus Type
            msc_quantity = buscar_informacao(df, i, "AT_MCP2_ROL_MSC_01_QTY")  # Coluna MSC Quantity
            merge_divert = buscar_informacao(df, i, "AT_MCP2_MRG_DIV_SEL")  # Coluna Merge/Divert
            merge_divert_angle = buscar_informacao(df, i, "AT_MCP2_MOD_MRG_ANG")  # Coluna Merge/Divert Angle
            alignment_angle = buscar_informacao(df, i, "AT_MCP2_FKT_ANG")  # Coluna Alignment Angle
            motor_position = buscar_informacao(df, i, "AT_MCP2_DRV_UNT_POS")  # Coluna Motor Position
            motor_manufacturer = buscar_informacao(df, i, "AT_MCP2_MOT_MNF")  # Coluna Motor Manufacturer
            framebed_type = buscar_informacao(df, i, "AT_MCP2_FRB_TYP")  # Coluna Framebed Type
            sword_quantity = buscar_informacao(df, i, "AT_MCP2_TRF_SWO_QTY")  # Coluna Sword Quantity
            cassetes_quantity = buscar_informacao(df, i, "AT_MCP2_CAS_QTY")  # Coluna Cassetes Quantity
            lower_conveyor_height = buscar_informacao(df, i, "AT_MCP2_LOW_LVL_CNV_HEI_M")  # Coluna Lower Conveyor Height
            higher_conveyor_height = buscar_informacao(df, i, "AT_MCP2_TOP_LVL_CNV_HEI_M")  # Coluna Higher Conveyor Height
            support_type = buscar_informacao(df, i, "AT_MCP2_SP_TYP_01")  # Coluna Support Type
            delivery_date = str(row[8]).strip() if len(row) > 8 else ""
            if delivery_date:
                delivery_date = delivery_date.replace('.', '/')
            dados.append([numero, codigo, quantidade, comprimento, between_frames, pitch, gear_ratio, power, voltage,
                          interface_type, sensor_type, tor, control_card, zone_length, electric_side, side_guide_left,
                          side_guide_right, bus_type, msc_quantity, merge_divert, merge_divert_angle, alignment_angle,
                          motor_position, motor_manufacturer, framebed_type, sword_quantity, cassetes_quantity,
                          lower_conveyor_height, higher_conveyor_height, support_type, data_criacao, delivery_date, customer])
    return dados

def buscar_comprimento(df, index, codigo):
    """
    Busca o comprimento relacionado a 'AT_MCP2_MOD_LEN_M' ou aplica regra especial para 'AT-RM8320-E2'.
    """
    try:
        if codigo == "AT-RM8320-E2":
            for i in range(index + 1, len(df)):  # Busca nas linhas seguintes
                row = df.iloc[i]
                for col in range(len(row)):
                    if "AT_MCP2_FKT_ANG" in str(row[col]):
                        angulo = row[col + 1] if col + 1 < len(row) else None
                        if angulo in [30, 45, 60, 90]:
                            return {30: 542, 45: 813, 60: 1084, 90: 1626}.get(angulo, 0)
        for i in range(index + 1, len(df)):  # Busca nas linhas seguintes
            row = df.iloc[i]
            for col in range(len(row)):
                if "AT_MCP2_MOD_LEN_M" in str(row[col]):
                    valor = row[col + 1] if col + 1 < len(row) else None
                    return int(float(valor)) if str(valor).replace('.', '', 1).isdigit() else 0
    except Exception as e:
        print(f"Erro ao buscar comprimento: {e}")
    return 0

def buscar_informacao(df, index, chave):
    """
    Busca uma informação específica pela chave no arquivo CSV.
    Se não encontrada, retorna 0.
    """
    try:
        for i in range(index + 1, len(df)):  # Busca nas linhas seguintes
            row = df.iloc[i]
            for col in range(len(row)):
                 if chave in str(row[col]):
                    valor = row[col + 1] if col + 1 < len(row) else None
                    # Substitui "NA" por "None"
                    if valor == "NA":
                        return "None"
                    return str(valor).strip() if valor else 0
    except Exception as e:
        print(f"Erro ao buscar a informação '{chave}': {e}")
    return 0

def atualizar_excel_existente(sales_orders_data, quotations_data, caminho_destino, nome_arquivo):
    """
    Atualiza os dados nas abas 'Sales Orders' e 'Quotations' em um arquivo Excel existente.
    """
    try:
        caminho_completo = os.path.join(caminho_destino, nome_arquivo)
        wb = load_workbook(caminho_completo)
        
        # Aba Sales Orders
        sales_orders_ws = wb["Sales Orders"]
        for item in sales_orders_data:
            sales_orders_ws.append(item)
        
        # Aba Quotations
        quotations_ws = wb["Quotations"]
        for item in quotations_data:
            quotations_ws.append(item)
        
        wb.save(caminho_completo)
        print(f"Arquivo Excel atualizado com sucesso em: {caminho_completo}")
    except Exception as e:
        print(f"Erro ao atualizar o arquivo Excel: {e}")

def verificar_novos_arquivos(arquivos_csv, caminho_destino, nome_arquivo):
    """
    Verifica se há novos arquivos CSV que ainda não estão no Excel existente.
    """
    caminho_completo = os.path.join(caminho_destino, nome_arquivo)
    wb = load_workbook(caminho_completo)
    
    sales_orders_ws = wb["Sales Orders"]
    quotations_ws = wb["Quotations"]
    
    # Obter os valores das primeiras colunas das abas Sales Orders e Quotations
    sales_orders_existentes = {row[0].value for row in sales_orders_ws.iter_rows(min_row=2, max_col=1) if row[0].value}
    quotations_existentes = {row[0].value for row in quotations_ws.iter_rows(min_row=2, max_col=1) if row[0].value}
    
    novos_arquivos = []
    for arquivo in arquivos_csv:
        tipo, df, numero, _, _ = processar_csv(arquivo)
        if tipo == 'Sales Orders' and numero not in sales_orders_existentes:
            novos_arquivos.append(arquivo)
        elif tipo == 'Quotations' and numero not in quotations_existentes:
            novos_arquivos.append(arquivo)
    
    return novos_arquivos

def executar_script():
    # Listar arquivos
    arquivos_csv = listar_arquivos(DIRETORIOS_ORIGEM)
    
    if arquivos_csv:
        # Verificar novos arquivos
        novos_arquivos = verificar_novos_arquivos(arquivos_csv, DIRETORIO_DESTINO, ARQUIVO_DESTINO)
        
        if novos_arquivos:
            sales_orders_data = []
            quotations_data = []
            
            # Processar novos arquivos CSV
            for arquivo in novos_arquivos:
                customer = extrair_cliente(os.path.basename(arquivo))
                tipo, df, numero, data_criacao, _ = processar_csv(arquivo)
                if tipo and df is not None:
                    dados = extrair_dados(df, tipo, numero, data_criacao, customer)
                    if tipo == 'Sales Orders':
                        sales_orders_data.extend(dados)
                    elif tipo == 'Quotations':
                        quotations_data.extend(dados)
            
            # Atualizar Excel existente
            atualizar_excel_existente(sales_orders_data, quotations_data, DIRETORIO_DESTINO, ARQUIVO_DESTINO)
        else:
            print("Nenhum arquivo novo encontrado.")
    else:
        print("Nenhum arquivo CSV encontrado nos diretórios de origem.")

if __name__ == "__main__":
    executar_script()